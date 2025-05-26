import pandas as pd
from woocommerce import API
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import logging
import jdatetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formataddr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- Logging Configuration ---
# Setting up basic logging for critical errors and important events
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Load Environment Variables ---
load_dotenv()

# --- WooCommerce API Configuration ---
WOO_BASE_URL = os.getenv("WOOCOMMERCE_BASE_URL")
WOO_CONSUMER_KEY = os.getenv("WOOCOMMERCE_CONSUMER_KEY")
WOO_CONSUMER_SECRET = os.getenv("WOOCOMMERCE_CONSUMER_SECRET")

if not all([WOO_BASE_URL, WOO_CONSUMER_KEY, WOO_CONSUMER_SECRET]):
    logger.critical("FATAL: Missing one or more WooCommerce API environment variables. Please check .env file.")
    exit(1)

try:
    wcapi = API(
        url=WOO_BASE_URL,
        consumer_key=WOO_CONSUMER_KEY,
        consumer_secret=WOO_CONSUMER_SECRET,
        version="wc/v3"
    )
except Exception as e:
    logger.critical(f"FATAL: Error configuring WooCommerce API: {e}")
    exit(1)

# --- Email Configuration ---
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER_TO = os.getenv("EMAIL_RECEIVER_TO")
EMAIL_RECEIVER_CC = os.getenv("EMAIL_RECEIVER_CC")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))

if not all([EMAIL_SENDER, EMAIL_PASSWORD, SMTP_SERVER, SMTP_PORT]):
    logger.warning("WARNING: One or more email environment variables (sender or server) are missing. Email sending might not function correctly.")

# --- Functions ---
def get_orders_from_yesterday():
    """Fetches all completed orders from the previous day from the WooCommerce API."""
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    all_orders = []
    page = 1
    per_page = 100

    try:
        while True:
            response_json = wcapi.get("orders", params={
                "status": "completed",
                "after": f"{yesterday}T00:00:00",
                "before": f"{yesterday}T23:59:59",
                "per_page": per_page,
                "page": page
            }).json()

            if not isinstance(response_json, list) or not response_json:
                break
            
            all_orders.extend(response_json)
            page += 1

    except Exception as e:
        logger.error(f"ERROR: Error fetching orders for {yesterday}: {e}")
        raise

    if not all_orders:
        logger.info(f"INFO: No completed orders found for {yesterday}.")
    
    return all_orders


def create_excel_report(orders_data):
    """Processes order data and generates an Excel report with styling."""
    processed_data = []
    for order in orders_data:
        try:
            item_names = [item['name'] for item in order.get('line_items', [])]

            item_quantities = []
            for item in order.get('line_items', []):
                quantity = item.get('quantity', 0)
                refunded_qty_for_this_item = 0
                for refund in order.get('refunds', []):
                    for refunded_item in refund.get('line_items', []):
                        if refunded_item.get('product_id') == item.get('product_id') and \
                           refunded_item.get('variation_id', 0) == item.get('variation_id', 0):
                            refunded_qty_for_this_item += refunded_item.get('qty', 0)

                item_quantities.append(str(quantity - refunded_qty_for_this_item))

            order_refund_total = sum(float(refund.get('total', 0)) for refund in order.get('refunds', []))

            created_datetime_obj = datetime.strptime(order['date_created'], '%Y-%m-%dT%H:%M:%S')
            jalali_date = jdatetime.datetime.fromtimestamp(created_datetime_obj.timestamp())
            formatted_jalali_date = jalali_date.strftime('%Y/%m/%d %H:%M:%S')

            custom_order_number = next(
                (meta['value'] for meta in order.get('meta_data', []) if meta['key'] == '_wc_order_number' or meta['key'] == '_order_number'),
                order.get('id')
            )

            order_row = {
                "شماره سفارش": custom_order_number,
                "تاریخ سفارش (شمسی)": formatted_jalali_date,
                "نام": order.get('billing', {}).get('first_name', ''),
                "نام خانوادگی": order.get('billing', {}).get('last_name', ''),
                "آدرس": f"{order.get('billing', {}).get('address_1', '')} {order.get('billing', {}).get('address_2', '')}".strip(),
                "شهر": order.get('billing', {}).get('city', ''),
                "کد پستی": order.get('billing', {}).get('postcode', ''),
                "تلفن": order.get('billing', {}).get('phone', ''),
                "عنوان روش پرداخت": order.get('payment_method_title', ''),
                "مبلغ تخفیف": float(order.get('discount_total', 0)),
                "مجموع مبلغ سفارش": float(order.get('total', 0)),
                "روش حمل و نقل": order.get('shipping_lines', [{}])[0].get('method_title', '') if order.get('shipping_lines') else '',
                "مبلغ حمل و نقل": float(order.get('shipping_lines', [{}])[0].get('total', 0)) if order.get('shipping_lines') else 0,
                "مبلغ استرداد کل سفارش": order_refund_total,
                "مجموع نهایی سفارش (پس از کسر استرداد)": float(order.get('total', 0)) - order_refund_total,
                "نام آیتم‌ها": "\n".join(item_names),
                "تعداد آیتم‌ها (- استرداد)": "\n".join(item_quantities),
                "مجموع هزینه آیتم‌ها": sum(float(item.get('total', 0)) for item in order.get('line_items', []))
            }
            processed_data.append(order_row)
        except Exception as e:
            logger.error(f"ERROR: Error processing order {order.get('id', 'N/A')}: {e}. This order was skipped.")
            continue

    df = pd.DataFrame(processed_data)
    
    excel_filename = f"WooCommerce_Orders_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    try:
        df.to_excel(excel_filename, index=False, engine='openpyxl')

        # --- Apply Excel Styling ---
        workbook = load_workbook(excel_filename)
        sheet = workbook.active

        header_fill = PatternFill(start_color="CCE0F0", end_color="CCE0F0", fill_type="solid")
        header_font = Font(bold=True)
        item_name_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
        wrap_text_alignment = Alignment(wrapText=True, vertical='top')

        for col_idx, cell in enumerate(sheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            max_length = 0
            column = get_column_letter(col_idx)
            for cell_in_col in sheet[column]:
                try:
                    if cell_in_col.value is not None:
                        content_lines = str(cell_in_col.value).split('\n')
                        for line in content_lines:
                            if len(line) > max_length:
                                max_length = len(line)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 70:
                adjusted_width = 70
            sheet.column_dimensions[column].width = adjusted_width

        sheet.freeze_panes = sheet['A2']

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in [get_column_letter(df.columns.get_loc("نام آیتم‌ها") + 1),
                                          get_column_letter(df.columns.get_loc("تعداد آیتم‌ها (- استرداد)") + 1),
                                          get_column_letter(df.columns.get_loc("آدرس") + 1)]:
                    cell.alignment = wrap_text_alignment

                if cell.column_letter == get_column_letter(df.columns.get_loc("نام آیتم‌ها") + 1):
                    cell.fill = item_name_fill
        
        workbook.save(excel_filename)
        logger.info(f"INFO: Excel file '{excel_filename}' generated and styled successfully.")
        return excel_filename
    except Exception as e:
        logger.error(f"ERROR: Error creating or styling Excel file '{excel_filename}': {e}")
        return None

def send_email_report(excel_file_path):
    """Sends an email with the generated Excel report as an attachment."""
    
    to_recipient = EMAIL_RECEIVER_TO if EMAIL_RECEIVER_TO else ""
    cc_recipient = EMAIL_RECEIVER_CC if EMAIL_RECEIVER_CC else ""

    if not EMAIL_SENDER or not EMAIL_PASSWORD or not SMTP_SERVER or not SMTP_PORT:
        logger.warning("WARNING: Email sending skipped due to missing sender/server credentials in .env.")
        return
    
    if not to_recipient and not cc_recipient:
        logger.warning("WARNING: Email sending skipped as no TO or CC recipients are specified in .env.")
        return

    msg = MIMEMultipart()
    msg['From'] = formataddr(('WooCommerce Report', EMAIL_SENDER))
    
    if to_recipient:
        msg['To'] = to_recipient
    if cc_recipient:
        msg['Cc'] = cc_recipient

    msg['Subject'] = f"گزارش سفارشات ووکامرس - {datetime.now().strftime('%Y-%m-%d')}"

    yesterday_datetime_obj = datetime.now() - timedelta(days=1)
    yesterday_jalali = jdatetime.datetime.fromtimestamp(yesterday_datetime_obj.timestamp()).strftime('%Y/%m/%d')
    
    body = f"با سلام،\n\nفایل اکسل گزارش سفارشات ووکامرس برای روز گذشته ({yesterday_jalali}) پیوست شده است.\n\nبا احترام - واحد انفورماتیک"
    msg.attach(MIMEText(body, 'plain'))

    if excel_file_path and os.path.exists(excel_file_path):
        try:
            with open(excel_file_path, 'rb') as f:
                attach = MIMEApplication(f.read(), _subtype="xlsx")
                attach.add_header('Content-Disposition', 'attachment', filename=os.path.basename(excel_file_path))
                msg.attach(attach)
            logger.info(f"INFO: Excel file '{os.path.basename(excel_file_path)}' successfully attached to email.")
        except Exception as e:
            logger.error(f"ERROR: Error attaching Excel file '{excel_file_path}': {e}. Email will be sent without attachment.")
    else:
        logger.warning(f"WARNING: Excel file '{excel_file_path}' not found or invalid path. Email will be sent without attachment.")

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            
            all_recipients = []
            if to_recipient:
                all_recipients.append(to_recipient)
            if cc_recipient:
                all_recipients.append(cc_recipient)

            server.send_message(msg, from_addr=EMAIL_SENDER, to_addrs=all_recipients)
            logger.info(f"INFO: Email successfully sent to '{to_recipient if to_recipient else 'N/A'}' and CC to '{cc_recipient if cc_recipient else 'N/A'}'.")
    except smtplib.SMTPAuthenticationError:
        logger.error("ERROR: SMTP Authentication Error: Check your email username and password in .env.")
    except smtplib.SMTPConnectError:
        logger.error(f"ERROR: SMTP Connection Error: Could not connect to '{SMTP_SERVER}' on port '{SMTP_PORT}'. Check server address, port, or network.")
    except Exception as e:
        logger.error(f"ERROR: An unexpected error occurred while sending email: {e}")

# --- Main Execution Flow ---
def main():
    """
    Main function to orchestrate fetching orders, generating report, and sending email.
    Handles critical errors during script execution.
    """
    logger.info("INFO: Script execution started.")
    try:
        orders = get_orders_from_yesterday()

        if orders:
            excel_file = create_excel_report(orders)
            if excel_file:
                send_email_report(excel_file)
            else:
                logger.error("ERROR: Excel file was not created, skipping email sending.")
        else:
            logger.info("INFO: No orders found for the previous day. No Excel report or email will be generated.")

    except Exception as e:
        logger.critical(f"CRITICAL: Script terminated due to a critical error: {e}", exc_info=True)

if __name__ == "__main__":
    main()