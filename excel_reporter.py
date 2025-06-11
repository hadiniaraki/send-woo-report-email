import pandas as pd
from datetime import datetime, timedelta
import os
import logging
import jdatetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import shutil

logger = logging.getLogger(__name__)

class ExcelReporter:
    """
    Handles the creation and styling of Excel reports from order data.
    """

    def _get_buyer_name(self, order):
        """Helper to get buyer name based on user type."""
        user_type = next((meta['value'] for meta in order.get('meta_data', []) if meta['key'] == '_user_type'), 'individual')
        if user_type == 'corporate':
            return order.get('billing', {}).get('company', '')
        else:
            first_name = order.get('billing', {}).get('first_name', '')
            last_name = order.get('billing', {}).get('last_name', '')
            return f"{first_name} {last_name}".strip()

    def create_excel_report(self, orders_data):
        """
        Processes order data and generates two Excel reports for the previous day.
        """
        processed_data = []

        yesterday_dt = datetime.now() - timedelta(days=1)
        yesterday_jalali_dt = jdatetime.datetime.fromgregorian(datetime=yesterday_dt)
        report_date_str = yesterday_jalali_dt.strftime('%Y-%m-%d')
        template_file = "tis.xlsx"
        templated_output_filename = None
        templated_output_filename_base = f"tis-{report_date_str}.xlsx"

        workbook_template = None
        sheet_body = None
        current_row_for_template = 0

        if os.path.exists(template_file):
            try:
                counter = 1
                unique_templated_output_filename = templated_output_filename_base
                while os.path.exists(unique_templated_output_filename):
                    unique_templated_output_filename = f"tis-{report_date_str}_{counter}.xlsx"
                    counter += 1
                templated_output_filename = unique_templated_output_filename

                shutil.copy(template_file, templated_output_filename)
                workbook_template = load_workbook(templated_output_filename)
                
                if "بدنه" in workbook_template.sheetnames:
                    sheet_body = workbook_template["بدنه"]
                    START_ROW_BODY = 2
                    current_row_for_template = START_ROW_BODY
                else:
                    logger.warning(f"WARNING: Sheet 'بدنه' not found in '{template_file}'. Templated report will not be generated.")
                    workbook_template = None
            except Exception as e:
                logger.error(f"ERROR: Could not copy or load template file '{template_file}': {e}.")
                workbook_template = None
        else:
            logger.error(f"ERROR: Template file '{template_file}' not found. Cannot create templated report.")


        for order in orders_data:
            try:
                # --- Common Data Extraction for both reports ---
                user_type = next((meta['value'] for meta in order.get('meta_data', []) if meta['key'] == '_user_type'), 'individual')
                
                company_name = order.get('billing', {}).get('company', '') if user_type == 'corporate' else ''
                national_id = next((meta['value'] for meta in order.get('meta_data', []) if meta['key'] == '_co_national_id'), '') if user_type == 'corporate' else ''
                register_id = next((meta['value'] for meta in order.get('meta_data', []) if meta['key'] == '_register_id'), '') if user_type == 'corporate' else ''

                item_names = []
                item_quantities = []
                item_prices_no_tax = []
                item_vat_amounts = []
                total_items_vat = 0.0
                total_items_price_no_tax = 0.0

                buyer_name = self._get_buyer_name(order) 

                # Process line items for both reports
                for item in order.get('line_items', []):
                    item_name = item['name']
                    quantity = item.get('quantity', 0)
                    item_total_price = float(item.get('total', 0))

                    refunded_qty_for_this_item = 0
                    for refund in order.get('refunds', []):
                        for refunded_item in refund.get('line_items', []):
                            if refunded_item.get('product_id') == item.get('product_id') and \
                               refunded_item.get('variation_id', 0) == item.get('variation_id', 0):
                                refunded_qty_for_this_item += refunded_item.get('qty', 0)
                    
                    effective_quantity = quantity - refunded_qty_for_this_item
                    
                    if effective_quantity <= 0:
                        continue

                    if item_total_price > 0:
                        price_no_tax_per_item = (item_total_price / 1.10) / effective_quantity
                        vat_per_item = (item_total_price / effective_quantity) - price_no_tax_per_item
                    else:
                        price_no_tax_per_item = 0.0
                        vat_per_item = 0.0
                    
                    item_names.append(item_name)
                    item_quantities.append(str(effective_quantity)) 
                    item_prices_no_tax.append(f"{price_no_tax_per_item * effective_quantity:,.0f}")
                    item_vat_amounts.append(f"{vat_per_item * effective_quantity:,.0f}")

                    total_items_vat += (vat_per_item * effective_quantity)
                    total_items_price_no_tax += (price_no_tax_per_item * effective_quantity)

                    # Populate the templated Excel file (only for individual customers)
                    if workbook_template and sheet_body and user_type == 'individual':
                        COL_DESCRIPTION, COL_QUANTITY, COL_UNIT, COL_UNIT_PRICE, COL_DISCOUNT, COL_VAT_RATE, COL_OTHER_TAX_SUBJECT = 3, 4, 5, 6, 10, 11, 12
                        try:
                            sheet_body.cell(row=current_row_for_template, column=COL_DESCRIPTION, value=item_name)
                            sheet_body.cell(row=current_row_for_template, column=COL_QUANTITY, value=effective_quantity)
                            sheet_body.cell(row=current_row_for_template, column=COL_UNIT, value="عدد")
                            sheet_body.cell(row=current_row_for_template, column=COL_UNIT_PRICE, value=round(price_no_tax_per_item)) 
                            sheet_body.cell(row=current_row_for_template, column=COL_DISCOUNT, value=0)
                            sheet_body.cell(row=current_row_for_template, column=COL_VAT_RATE, value=10)
                            sheet_body.cell(row=current_row_for_template, column=COL_OTHER_TAX_SUBJECT, value=buyer_name)
                            current_row_for_template += 1
                        except Exception as e:
                            logger.error(f"ERROR: Error writing item '{item_name}' of order {order.get('id', 'N/A')} to templated report: {e}")

                order_refund_total = sum(float(refund.get('total', 0)) for refund in order.get('refunds', []))
                created_datetime_obj = datetime.strptime(order['date_created'], '%Y-%m-%dT%H:%M:%S')
                jalali_date = jdatetime.datetime.fromgregorian(datetime=created_datetime_obj)
                formatted_jalali_date = jalali_date.strftime('%Y/%m/%d %H:%M:%S')

                custom_order_number = next((meta['value'] for meta in order.get('meta_data', []) if meta['key'] in ['_wc_order_number', '_order_number']), order.get('id'))

                order_row = {
                    "شماره سفارش": custom_order_number, "تاریخ سفارش (شمسی)": formatted_jalali_date,
                    "نام": order.get('billing', {}).get('first_name', ''), "نام خانوادگی": order.get('billing', {}).get('last_name', ''),
                    "نام شرکت": company_name, "شناسه ملی": national_id, "شماره ثبت": register_id,
                    "آدرس": f"{order.get('billing', {}).get('address_1', '')} {order.get('billing', {}).get('address_2', '')}".strip(),
                    "شهر": order.get('billing', {}).get('city', ''), "کد پستی": order.get('billing', {}).get('postcode', ''),
                    "تلفن": order.get('billing', {}).get('phone', ''), "عنوان روش پرداخت": order.get('payment_method_title', ''),
                    "مبلغ تخفیف": float(order.get('discount_total', 0)),
                    "مجموع مبلغ سفارش (با مالیات)": float(order.get('total', 0)),
                    "مجموع نهایی سفارش (بدون مالیات)": total_items_price_no_tax,
                    "مجموع مالیات بر ارزش افزوده": total_items_vat,
                    "روش حمل و نقل": order.get('shipping_lines', [{}])[0].get('method_title', '') if order.get('shipping_lines') else '',
                    "مبلغ حمل و نقل": float(order.get('shipping_lines', [{}])[0].get('total', 0)) if order.get('shipping_lines') else 0,
                    "مبلغ استرداد کل سفارش": order_refund_total,
                    "مجموع نهایی سفارش (پس از کسر استرداد و با مالیات)": float(order.get('total', 0)) - order_refund_total,
                    "نام آیتم‌ها": "\n".join(item_names), "تعداد آیتم‌ها (- استرداد)": "\n".join(item_quantities),
                    "قیمت واحد آیتم (بدون مالیات)": "\n".join(item_prices_no_tax), "مالیات بر ارزش افزوده آیتم": "\n".join(item_vat_amounts),
                    "مجموع هزینه آیتم‌ها (با مالیات)": sum(float(item.get('total', 0)) for item in order.get('line_items', []))
                }
                processed_data.append(order_row)
            except Exception as e:
                logger.error(f"ERROR: Error processing order {order.get('id', 'N/A')}: {e}.")
                continue

        # Save the templated workbook
        if workbook_template and templated_output_filename:
            try:
                workbook_template.save(templated_output_filename)
                logger.info(f"INFO: Templated Excel file '{templated_output_filename}' generated successfully.")
            except Exception as e:
                logger.error(f"ERROR: Error saving templated Excel file '{templated_output_filename}': {e}")
                templated_output_filename = None
        else:
            templated_output_filename = None

        # --- Generate the main comprehensive report ---
        if not processed_data:
             return None, []
             
        df = pd.DataFrame(processed_data)
        if "تاریخ سفارش (شمسی)" in df.columns:
            df = df.sort_values(by="تاریخ سفارش (شمسی)", ascending=True)

        # ===> [تغییر] استفاده از نام و تاریخ جدید برای فایل اصلی
        main_excel_filename = f"Orders_سایت_{report_date_str}.xlsx"
        try:
            df.to_excel(main_excel_filename, index=False, engine='openpyxl')

            # --- Apply Excel Styling to main report ---
            workbook = load_workbook(main_excel_filename)
            sheet = workbook.active
            header_fill = PatternFill(start_color="CCE0F0", end_color="CCE0F0", fill_type="solid")
            header_font = Font(bold=True)
            item_name_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
            wrap_text_alignment = Alignment(wrapText=True, vertical='top')

            for col_idx, cell in enumerate(sheet[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                max_length = max(len(str(cell_in_col.value).split('\n')[0]) for cell_in_col in sheet[get_column_letter(col_idx)] if cell_in_col.value is not None)
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(adjusted_width, 70)

            sheet.freeze_panes = sheet['A2']
            
            wrap_columns = ["نام آیتم‌ها", "تعداد آیتم‌ها (- استرداد)", "قیمت واحد آیتم (بدون مالیات)", "مالیات بر ارزش افزوده آیتم", "آدرس"]
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    col_title = sheet.cell(row=1, column=cell.column).value
                    if col_title in wrap_columns:
                        cell.alignment = wrap_text_alignment
                    if col_title == "نام آیتم‌ها":
                        cell.fill = item_name_fill

            workbook.save(main_excel_filename)
            logger.info(f"INFO: Main Excel file '{main_excel_filename}' generated and styled successfully.")
            
            return main_excel_filename, [templated_output_filename] if templated_output_filename and os.path.exists(templated_output_filename) else []

        except Exception as e:
            logger.error(f"ERROR: Error creating or styling main Excel file '{main_excel_filename}': {e}")
            return None, [templated_output_filename] if templated_output_filename and os.path.exists(templated_output_filename) else []